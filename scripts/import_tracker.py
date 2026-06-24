"""Read-only tracker inspection utility.

This script never writes back to the source workbook. It exists for Stage 0
inspection and later controlled imports.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SHEETS = ("Company Watchlist", "Pipeline", "History")


def sheet_summary(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    summaries: list[dict[str, Any]] = []

    for name in workbook.sheetnames:
        sheet = workbook[name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        non_empty_rows = 0
        observed_max_columns = 0

        for row in sheet.iter_rows(values_only=True):
            if any(value is not None for value in row):
                non_empty_rows += 1
                observed_max_columns = max(
                    observed_max_columns,
                    max((index for index, value in enumerate(row, start=1) if value is not None), default=0),
                )

        summaries.append(
            {
                "name": name,
                "non_empty_rows": non_empty_rows,
                "observed_max_columns": observed_max_columns,
                "headers": headers,
                "stage0_required": name in DEFAULT_SHEETS,
            }
        )

    return {"workbook": str(path), "sheets": summaries}


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect the job tracker workbook read-only.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out", type=Path, help="Optional JSON output path.")
    args = parser.parse_args()

    summary = sheet_summary(args.workbook)
    text = json.dumps(summary, indent=2, default=str)

    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(text + "\n", encoding="utf-8")
    else:
        print(text)


if __name__ == "__main__":
    main()
